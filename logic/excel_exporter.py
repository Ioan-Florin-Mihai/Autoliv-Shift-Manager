from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from logic.constants import HOURS_12_COLOR
from logic.schedule_store import DAY_NAMES, DEPARTMENT_COLORS, SHIFTS

THIN_SIDE = Side(style="thin", color="8A8A8A")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BLUE_FILL = PatternFill("solid", fgColor="4F6FB5")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
BRAND_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
DEPARTMENT_FONT = Font(name="Calibri", size=9, bold=True)
HEADER_FONT = Font(name="Calibri", size=9, bold=True)
BODY_FONT = Font(name="Calibri", size=8)
RED_FONT = Font(name="Calibri", size=8, color="C0392B")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
VERTICAL_CENTER = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

DISPLAY_DAY_NAMES = {
    "Luni": "Luni",
    "Marti": "Marți",
    "Miercuri": "Miercuri",
    "Joi": "Joi",
    "Vineri": "Vineri",
    "Sambata": "Sâmbătă",
    "Duminica": "Duminică",
}

DEPARTMENT_FILL_OVERRIDES = {
    "sef schimb": "5B9BD5",
    "receptii": "5B9BD5",
    "livrari": "A9D18E",
    "etichetare scanare": "C9B0D9",
    "retragere finite": "D9A35F",
    "ambalaje": "D99694",
    "balotare ambalare": "BFBFBF",
}

SHEET_DEPARTMENTS: dict[str, list[tuple[str, str]]] = {
    "Bucle": [
        ("BUCLA RA+RB", "BUCLA RA + RB"),
        ("BUCLA TA+TB", "BUCLA TA + TB"),
        ("BUCLA 02", "BUCLA 02"),
        ("BUCLA 03", "BUCLA 03"),
        ("BUCLA 04", "BUCLA 04"),
        ("BUCLA 05", "BUCLA 05 + 07"),
        ("Ambalaje", "Ambalaje"),
    ],
    "Magazie": [
        ("Sef schimb", "Sef Schimb"),
        ("Receptii", "Receptii"),
        ("Livrari", "Livrari"),
        ("Etichetare scanare", "Etichetare / Scanare"),
        ("Retragere finite", "Retragere finite"),
        ("Balotare ambalare", "Baloare / Asamblare"),
    ],
}

SHEET_ROW_LAYOUT = {
    "Bucle": {"block_header": 25, "shift": 34},
    "Magazie": {"block_header": 30, "shift": 43},
}


def _parse_week_start(week_start: str) -> date | None:
    try:
        return date.fromisoformat(str(week_start))
    except ValueError:
        return None


def _day_date_labels(week_start: str) -> list[tuple[str, str]]:
    base = _parse_week_start(week_start)
    if base is None:
        return [(day, "") for day in DAY_NAMES]

    def fmt(current_date: date) -> str:
        month = current_date.strftime("%b").lower()
        month_map = {
            "jan": "ian",
            "feb": "feb",
            "mar": "mar",
            "apr": "apr",
            "may": "mai",
            "jun": "iun",
            "jul": "iul",
            "aug": "aug",
            "sep": "sep",
            "oct": "oct",
            "nov": "nov",
            "dec": "dec",
        }
        return f"{current_date.day}-{month_map.get(month, month)}.-{current_date.strftime('%y')}"

    return [(day_name, fmt(base + timedelta(days=offset))) for offset, day_name in enumerate(DAY_NAMES)]


def _iter_export_departments(modes: dict) -> list[tuple[str, dict, str]]:
    result: list[tuple[str, dict, str]] = []
    if not isinstance(modes, dict):
        return result
    for mode_name, mode_record in modes.items():
        if not isinstance(mode_record, dict):
            continue
        departments = mode_record.get("departments", [])
        if not isinstance(departments, list):
            continue
        for department in departments:
            if isinstance(department, str) and department.strip():
                result.append((str(mode_name), mode_record, department))
    return result


def _safe_cell(mode_record: dict, department: str, day_name: str, shift: str) -> dict:
    try:
        cell = mode_record["schedule"][department][day_name][shift]
    except (KeyError, TypeError):
        return {"employees": [], "colors": {}}
    if not isinstance(cell, dict):
        return {"employees": [], "colors": {}}
    employees = cell.get("employees", [])
    colors = cell.get("colors", {})
    return {
        "employees": employees if isinstance(employees, list) else [],
        "colors": colors if isinstance(colors, dict) else {},
    }


def _hours_label_for_employee(cell: dict, employee: str) -> str:
    colors = cell.get("colors", {}) if isinstance(cell, dict) else {}
    if not isinstance(colors, dict):
        return "8h"
    for key, value in colors.items():
        if isinstance(key, str) and key.casefold() == employee.casefold():
            if str(value or "").strip().upper().lstrip("#") == HOURS_12_COLOR:
                return "12h"
            return "8h"
    return "8h"


def _department_fill(department: str) -> PatternFill:
    normalized = " ".join(str(department or "").split()).casefold()
    if normalized.startswith("bucla"):
        color = "D9A35F"
    else:
        color = DEPARTMENT_FILL_OVERRIDES.get(normalized, DEPARTMENT_COLORS.get(department, "D9D9D9"))
    color = str(color).strip().lstrip("#") or "D9D9D9"
    return PatternFill("solid", fgColor=color)


def _title_week_label(week_record: dict) -> str:
    label = str(week_record.get("week_label", "") or "")
    return label.replace("Saptamana", "saptamana").replace("Săptămâna", "saptamana")


def _write_bordered_cell(ws, row: int, col: int, value="", *, fill=None, font=None, alignment=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.fill = fill or WHITE_FILL
    cell.font = font or BODY_FONT
    cell.alignment = alignment or CENTER


def _sheet_title(mode_name: str, week_label: str) -> str:
    base = mode_name[:20] if mode_name else "Planificare"
    return base or week_label[:20] or "Planificare"


def _configure_sheet(ws, title: str) -> None:
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True


def _render_mode_sheet(
    ws,
    *,
    week_record: dict,
    mode_name: str,
    mode_record: dict,
    departments: list[tuple[str, str]],
) -> None:
    day_labels = _day_date_labels(str(week_record.get("week_start") or ""))
    last_col = 9

    widths = {
        "A": 6.5,
        "B": 11.0,
        "C": 24.0,
        "D": 24.0,
        "E": 24.0,
        "F": 24.0,
        "G": 24.0,
        "H": 23.0,
        "I": 23.0,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 8

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=f"Planificare magazie : {_title_week_label(week_record)}")
    title_cell.fill = BLUE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.border = THIN_BORDER
    for col in range(2, 9):
        cell = ws.cell(row=1, column=col)
        cell.fill = BLUE_FILL
        cell.border = THIN_BORDER

    brand_cell = ws.cell(row=1, column=9, value="Autoliv")
    brand_cell.fill = BLUE_FILL
    brand_cell.font = BRAND_FONT
    brand_cell.alignment = CENTER
    brand_cell.border = THIN_BORDER

    for col in range(1, last_col + 1):
        ws.cell(row=2, column=col).fill = WHITE_FILL

    row_layout = SHEET_ROW_LAYOUT.get(mode_name, SHEET_ROW_LAYOUT["Magazie"])

    current_row = 3
    for department_key, department_label in departments:
        header_row = current_row
        first_shift_row = header_row + 1
        last_shift_row = header_row + len(SHIFTS)
        ws.row_dimensions[header_row].height = row_layout["block_header"]
        for row in range(first_shift_row, last_shift_row + 1):
            ws.row_dimensions[row].height = row_layout["shift"]

        ws.merge_cells(start_row=header_row, start_column=1, end_row=last_shift_row, end_column=1)
        dept_cell = ws.cell(row=header_row, column=1, value=department_label)
        dept_cell.fill = _department_fill(department_key)
        dept_cell.font = DEPARTMENT_FONT
        dept_cell.alignment = VERTICAL_CENTER
        dept_cell.border = THIN_BORDER

        _write_bordered_cell(ws, header_row, 2, "Schimbul", fill=HEADER_FILL, font=HEADER_FONT)
        for day_index, (day_name, date_label) in enumerate(day_labels):
            current_col = 3 + day_index
            display_day = DISPLAY_DAY_NAMES.get(day_name, day_name)
            header_value = f"{display_day}\n{date_label}".strip()
            _write_bordered_cell(ws, header_row, current_col, header_value, fill=HEADER_FILL, font=HEADER_FONT)

        for shift_index, shift in enumerate(SHIFTS):
            row = first_shift_row + shift_index
            _write_bordered_cell(ws, row, 2, shift, fill=WHITE_FILL, font=HEADER_FONT)
            for day_index, (day_name, _date_label) in enumerate(day_labels):
                current_col = 3 + day_index
                cell = _safe_cell(mode_record, department_key, day_name, shift)
                employees = [employee.strip() for employee in cell["employees"] if isinstance(employee, str) and employee.strip()]
                text_lines = [f"{employee}/{_hours_label_for_employee(cell, employee)}" for employee in employees] or ["-"]
                body_value = "\n".join(text_lines)
                is_special = any(_hours_label_for_employee(cell, employee) == "12h" for employee in employees)
                fill = HEADER_FILL if day_name in {"Sambata", "Duminica"} else WHITE_FILL
                _write_bordered_cell(
                    ws,
                    row,
                    current_col,
                    body_value,
                    fill=fill,
                    font=RED_FONT if is_special else BODY_FONT,
                    alignment=TOP_LEFT,
                )

        for row in range(header_row, last_shift_row + 1):
            dept_border_cell = ws.cell(row=row, column=1)
            dept_border_cell.border = THIN_BORDER
            dept_border_cell.fill = _department_fill(department_key)

        current_row = last_shift_row + 1

    ws.freeze_panes = ws["C3"]
    last_row = max(current_row - 1, 1)
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"


def export_full_week_plan_excel(*, week_record: dict, output_path: str | Path) -> None:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    modes = week_record.get("modes", {}) if isinstance(week_record, dict) else {}
    if not isinstance(modes, dict) or not modes:
        ws = workbook.create_sheet("Planificare")
        _configure_sheet(ws, "Planificare")
        ws["A1"] = "Nu exista posturi de exportat."
    else:
        for mode_name in ("Bucle", "Magazie"):
            mode_record = modes.get(mode_name, {})
            if not isinstance(mode_record, dict):
                mode_record = {}
            ws = workbook.create_sheet(mode_name)
            _configure_sheet(ws, mode_name)
            _render_mode_sheet(
                ws,
                week_record=week_record,
                mode_name=mode_name,
                mode_record=mode_record,
                departments=SHEET_DEPARTMENTS[mode_name],
            )

    workbook.save(output_path)
