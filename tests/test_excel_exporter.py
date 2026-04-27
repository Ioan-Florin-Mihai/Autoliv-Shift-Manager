from datetime import date

from openpyxl import load_workbook

from logic.schedule_store import DAY_NAMES, ScheduleStore

EXPECTED_DAY_LABELS = {
    "Luni": "Luni",
    "Marti": "Marți",
    "Miercuri": "Miercuri",
    "Joi": "Joi",
    "Vineri": "Vineri",
    "Sambata": "Sâmbătă",
    "Duminica": "Duminică",
}


def test_export_full_week_plan_excel_creates_print_ready_workbook(tmp_path, monkeypatch):
    import logic.schedule_store as store_module

    monkeypatch.setattr(store_module, "SCHEDULE_PATH", tmp_path / "schedule_data.json")
    monkeypatch.setattr(store_module, "BACKUP_DIR", tmp_path / "backups")

    store = ScheduleStore()
    week = store.get_or_create_week(date(2026, 4, 20))
    week["modes"]["Magazie"]["schedule"]["Sef schimb"]["Luni"]["Sch1"]["employees"] = ["Munteanu M"]
    week["modes"]["Magazie"]["schedule"]["Sef schimb"]["Luni"]["Sch1"]["colors"]["Munteanu M"] = "#C0392B"

    from logic.excel_exporter import export_full_week_plan_excel

    out = tmp_path / "planificare.xlsx"
    export_full_week_plan_excel(week_record=week, output_path=out)

    assert out.exists()

    workbook = load_workbook(out)
    assert workbook.sheetnames == ["Bucle", "Magazie"]

    sheet = workbook["Magazie"]
    assert sheet.page_setup.orientation == "landscape"
    assert int(sheet.page_setup.paperSize) == 8
    assert int(sheet.page_setup.fitToWidth) == 1
    assert int(sheet.page_setup.fitToHeight) == 1
    assert sheet.print_options.horizontalCentered is True
    assert sheet.print_options.verticalCentered is True
    assert sheet.print_area == "'Magazie'!$A$1:$I$26"
    assert sheet.column_dimensions["A"].width == 6.5
    assert sheet.column_dimensions["B"].width == 11.0
    assert sheet.column_dimensions["C"].width == 24.0
    assert sheet.column_dimensions["I"].width == 23.0
    assert sheet.row_dimensions[1].height == 34
    assert sheet.row_dimensions[3].height == 30
    assert sheet.row_dimensions[4].height == 43
    assert sheet["A1"].value == "Planificare magazie : saptamana 17"
    assert sheet["I1"].value == "Autoliv"
    assert sheet["A3"].value == "Sef Schimb"
    assert sheet["B3"].value == "Schimbul"
    assert "Luni" in str(sheet["C3"].value)
    assert "20-apr.-26" in str(sheet["C3"].value)
    assert "Munteanu M/12h" in str(sheet["C4"].value)
    assert "A1:H1" in {str(merged_range) for merged_range in sheet.merged_cells.ranges}
    assert "A3:A6" in {str(merged_range) for merged_range in sheet.merged_cells.ranges}
    assert not sheet.row_breaks.brk
    assert not sheet.col_breaks.brk


def test_excel_export_includes_all_departments_for_mode(tmp_path, monkeypatch):
    import logic.schedule_store as store_module

    monkeypatch.setattr(store_module, "SCHEDULE_PATH", tmp_path / "schedule_data.json")
    monkeypatch.setattr(store_module, "BACKUP_DIR", tmp_path / "backups")

    store = ScheduleStore()
    week = store.get_or_create_week(date(2026, 4, 20))

    from logic.excel_exporter import export_full_week_plan_excel

    out = tmp_path / "planificare.xlsx"
    export_full_week_plan_excel(week_record=week, output_path=out)

    workbook = load_workbook(out)
    sheet = workbook["Bucle"]
    assert sheet.page_setup.orientation == "landscape"
    assert int(sheet.page_setup.paperSize) == 8
    assert int(sheet.page_setup.fitToWidth) == 1
    assert int(sheet.page_setup.fitToHeight) == 1
    assert sheet.print_options.horizontalCentered is True
    assert sheet.print_options.verticalCentered is True
    assert sheet.print_area == "'Bucle'!$A$1:$I$30"
    assert sheet.row_dimensions[3].height == 25
    assert sheet.row_dimensions[4].height == 34
    assert not sheet.row_breaks.brk
    assert not sheet.col_breaks.brk
    expected_ranges = {str(merged_range) for merged_range in sheet.merged_cells.ranges}
    expected_departments = [
        "BUCLA RA + RB",
        "BUCLA TA + TB",
        "BUCLA 02",
        "BUCLA 03",
        "BUCLA 04",
        "BUCLA 05 + 07",
        "Ambalaje",
    ]
    for index, department in enumerate(expected_departments):
        block_start = 3 + (index * 4)
        assert sheet.cell(row=block_start, column=1).value == department
        assert f"A{block_start}:A{block_start + 3}" in expected_ranges
        assert sheet.cell(row=block_start, column=2).value == "Schimbul"
        for day_index, day_name in enumerate(DAY_NAMES):
            assert EXPECTED_DAY_LABELS[day_name] in str(sheet.cell(row=block_start, column=3 + day_index).value)
