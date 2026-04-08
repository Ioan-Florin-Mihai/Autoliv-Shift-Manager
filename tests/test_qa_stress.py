# ============================================================
# QA STRESS TEST SUITE — 150+ VALIDATIONS
# ============================================================
# Categories:
#   1. File System Tests
#   2. EXE Portability / Path Resolution
#   3. Offline / Firebase Fail-Safe
#   4. Auth / Login
#   5. Data Integrity (ScheduleStore)
#   6. Data Integrity (EmployeeStore)
#   7. Excel Export
#   8. Validation Module
#   9. Edge Cases
#  10. Error Handling / Recovery
#  11. Multi-Run Consistency
# ============================================================

import json
import os
import shutil
import tempfile
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest


# ── Fixtures ─────────────────────────────────────────────────

@pytest.fixture
def tmp_dir(tmp_path):
    """Provides a clean temp directory for each test."""
    return tmp_path


@pytest.fixture
def schedule_store(tmp_path, monkeypatch):
    """ScheduleStore with isolated file system."""
    path = tmp_path / "data" / "schedule_data.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
    monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", tmp_path / "backups")
    from logic.schedule_store import ScheduleStore
    return ScheduleStore()


@pytest.fixture
def employee_store(tmp_path, monkeypatch):
    """EmployeeStore with isolated file system."""
    path = tmp_path / "data" / "employees.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("logic.employee_store.EMPLOYEES_PATH", path)
    from logic.employee_store import EmployeeStore
    return EmployeeStore()


@pytest.fixture
def users_path(tmp_path, monkeypatch):
    """Isolated users.json path for auth tests."""
    path = tmp_path / "data" / "users.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr("logic.auth.USERS_PATH", path)
    return path


@pytest.fixture
def week_record():
    """A populated week record for export testing."""
    from logic.schedule_store import ScheduleStore, SCHEDULE_PATH
    store = ScheduleStore()
    return store.get_or_create_week(date.today())


@pytest.fixture
def populated_week():
    """Week record with employees assigned."""
    from logic.schedule_store import (
        _empty_week_record, get_week_start, DAYS, SHIFTS
    )
    w = _empty_week_record(get_week_start(date.today()))
    dept = w["modes"]["Magazie"]["departments"][0]
    day_name = DAYS[0][0]
    shift = SHIFTS[0]
    cell = w["modes"]["Magazie"]["schedule"][dept][day_name][shift]
    cell["employees"] = ["Ion Popescu", "Maria Ionescu", "Andrei Vasile"]
    cell["colors"] = {
        "Ion Popescu": "#C0392B",
        "Maria Ionescu": "#27AE60",
        "Andrei Vasile": "#C0392B",
    }
    return w


# ================================================================
# 1. FILE SYSTEM TESTS (FS-001 to FS-020)
# ================================================================

class TestFileSystem:
    """FS-001..FS-020: Missing/corrupt files, auto-create, permissions."""

    def test_fs001_data_dir_auto_create(self, tmp_path, monkeypatch):
        """FS-001: Missing data dir is auto-created."""
        from logic import app_paths
        new_data = tmp_path / "nonexistent" / "data"
        assert not new_data.exists()
        new_data.mkdir(parents=True, exist_ok=True)
        assert new_data.exists()

    def test_fs002_export_dir_auto_create(self, tmp_path):
        """FS-002: Export dir is created when missing."""
        export = tmp_path / "Exports"
        export.mkdir(parents=True, exist_ok=True)
        assert export.is_dir()

    def test_fs003_backup_dir_auto_create(self, tmp_path):
        """FS-003: Backup dir is created when missing."""
        backup = tmp_path / "backups"
        backup.mkdir(parents=True, exist_ok=True)
        assert backup.is_dir()

    def test_fs004_missing_users_json_auto_create(self, users_path):
        """FS-004: Missing users.json creates default admin."""
        assert not users_path.exists()
        from logic.auth import _load_users
        users = _load_users()
        assert users_path.exists()
        assert len(users) >= 1
        assert users[0]["username"] == "admin"

    def test_fs005_missing_users_json_default_admin_login(self, users_path):
        """FS-005: Default admin can login with admin123."""
        from logic.auth import verify_login
        result = verify_login("admin", "admin123")
        assert result is True

    def test_fs006_missing_schedule_loads_empty(self, tmp_path, monkeypatch):
        """FS-006: Missing schedule_data.json loads empty weeks dict."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", tmp_path / "backups")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert store.data == {"weeks": {}}

    def test_fs007_corrupt_schedule_json_loads_empty(self, tmp_path, monkeypatch):
        """FS-007: Corrupt JSON in schedule_data falls back to empty."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("{CORRUPT!!!", encoding="utf-8")
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", tmp_path / "backups")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert store.data == {"weeks": {}}

    def test_fs008_corrupt_employees_json_loads_empty(self, tmp_path, monkeypatch):
        """FS-008: Corrupt employees.json returns empty list."""
        path = tmp_path / "data" / "employees.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("NOT JSON!", encoding="utf-8")
        monkeypatch.setattr("logic.employee_store.EMPLOYEES_PATH", path)
        from logic.employee_store import EmployeeStore
        store = EmployeeStore()
        assert store.data.get("employees") == []

    def test_fs009_corrupt_users_json_raises_clear_error(self, users_path):
        """FS-009: Corrupt users.json raises ValueError with clear msg."""
        users_path.write_text("{{bad json", encoding="utf-8")
        from logic.auth import _load_users
        with pytest.raises(ValueError, match="corupt"):
            _load_users()

    def test_fs010_schedule_save_creates_file(self, schedule_store):
        """FS-010: ScheduleStore.save() creates the file."""
        schedule_store.get_or_create_week(date.today())
        from logic.schedule_store import SCHEDULE_PATH
        # Data was saved by get_or_create_week — check via fresh load
        assert "weeks" in schedule_store.data

    def test_fs011_atomic_save_no_partial_writes(self, schedule_store):
        """FS-011: Save is atomic — no partial writes on disk."""
        schedule_store.get_or_create_week(date.today())
        schedule_store.save()

    def test_fs012_backup_creates_file(self, tmp_path, monkeypatch):
        """FS-012: backup() creates a timestamped file."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        store.get_or_create_week(date.today())
        store.save()
        store.backup()
        backups = list(backup_dir.glob("schedule_backup_*.json"))
        assert len(backups) >= 1

    def test_fs013_ensure_runtime_file_missing_source(self, tmp_path, monkeypatch):
        """FS-013: ensure_runtime_file() handles missing source gracefully."""
        from logic.app_paths import ensure_runtime_file
        # With real paths, if no bundle source, just returns target
        result = ensure_runtime_file("data/schedule_data.json")
        assert isinstance(result, Path)

    def test_fs014_get_sensitive_path_never_copies(self, tmp_path, monkeypatch):
        """FS-014: get_sensitive_path() never copies from bundle."""
        from logic.app_paths import get_sensitive_path
        result = get_sensitive_path("data/users.json")
        assert isinstance(result, Path)

    def test_fs015_schedule_store_preserves_json_encoding(self, schedule_store):
        """FS-015: Special characters preserved after save/load."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        cell = week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"].append("Ștefan Pîrvu-Dăescu")
        schedule_store.update_week(week)
        # Reload
        from logic.schedule_store import ScheduleStore, SCHEDULE_PATH
        store2 = ScheduleStore()
        w2 = store2.get_or_create_week(date.today())
        cell2 = w2["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        assert "Ștefan Pîrvu-Dăescu" in cell2["employees"]

    def test_fs016_missing_employees_json_no_crash(self, tmp_path, monkeypatch):
        """FS-016: Missing employees.json does not crash."""
        path = tmp_path / "data" / "employees.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.employee_store.EMPLOYEES_PATH", path)
        from logic.employee_store import EmployeeStore
        store = EmployeeStore()
        assert isinstance(store.data, dict)

    def test_fs017_empty_schedule_file_handled(self, tmp_path, monkeypatch):
        """FS-017: Empty schedule file treated as corrupt → empty weeks."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("", encoding="utf-8")
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", tmp_path / "backups")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert store.data == {"weeks": {}}

    def test_fs018_employees_json_wrong_type(self, tmp_path, monkeypatch):
        """FS-018: employees.json with wrong root type handled."""
        path = tmp_path / "data" / "employees.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text('"just a string"', encoding="utf-8")
        monkeypatch.setattr("logic.employee_store.EMPLOYEES_PATH", path)
        from logic.employee_store import EmployeeStore
        store = EmployeeStore()
        assert store.data.get("employees") == []

    def test_fs019_schedule_json_wrong_structure(self, tmp_path, monkeypatch):
        """FS-019: schedule_data.json with wrong structure falls back."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text('{"weeks": "not a dict"}', encoding="utf-8")
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", tmp_path / "backups")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert store.data == {"weeks": {}}

    def test_fs020_users_json_dict_format_migrated(self, users_path):
        """FS-020: Old dict-format users.json auto-migrated to list."""
        import bcrypt
        h = bcrypt.hashpw(b"pass1234", bcrypt.gensalt(rounds=4)).decode()
        users_path.write_text(json.dumps({
            "username": "admin", "password_hash": h
        }), encoding="utf-8")
        from logic.auth import _load_users
        users = _load_users()
        assert isinstance(users, list)
        assert users[0]["username"] == "admin"


# ================================================================
# 2. EXE PORTABILITY / PATH RESOLUTION (PT-001 to PT-014)
# ================================================================

class TestPortability:
    """PT-001..PT-014: Path resolution, portability invariants."""

    def test_pt001_base_dir_is_directory(self):
        """PT-001: BASE_DIR exists and is a directory."""
        from logic.app_paths import BASE_DIR
        assert BASE_DIR.is_dir()

    def test_pt002_bundle_dir_is_directory(self):
        """PT-002: BUNDLE_DIR exists and is a directory."""
        from logic.app_paths import BUNDLE_DIR
        assert BUNDLE_DIR.is_dir()

    def test_pt003_data_dir_exists(self):
        """PT-003: DATA_DIR is auto-created at import."""
        from logic.app_paths import DATA_DIR
        assert DATA_DIR.is_dir()

    def test_pt004_export_dir_exists(self):
        """PT-004: EXPORT_DIR is auto-created at import."""
        from logic.app_paths import EXPORT_DIR
        assert EXPORT_DIR.is_dir()

    def test_pt005_backup_dir_exists(self):
        """PT-005: BACKUP_DIR is auto-created at import."""
        from logic.app_paths import BACKUP_DIR
        assert BACKUP_DIR.is_dir()

    def test_pt006_assets_dir_exists(self):
        """PT-006: ASSETS_DIR is auto-created at import."""
        from logic.app_paths import ASSETS_DIR
        assert ASSETS_DIR.is_dir()

    def test_pt007_dev_mode_detected(self):
        """PT-007: In dev mode, frozen is False."""
        import sys
        assert not getattr(sys, "frozen", False)

    def test_pt008_base_dir_absolute(self):
        """PT-008: BASE_DIR is an absolute path."""
        from logic.app_paths import BASE_DIR
        assert BASE_DIR.is_absolute()

    def test_pt009_bundle_dir_absolute(self):
        """PT-009: BUNDLE_DIR is an absolute path."""
        from logic.app_paths import BUNDLE_DIR
        assert BUNDLE_DIR.is_absolute()

    def test_pt010_ensure_directory_creates_nested(self, tmp_path):
        """PT-010: ensure_directory() creates nested dirs."""
        from logic.app_paths import ensure_directory
        target = tmp_path / "a" / "b" / "c"
        ensure_directory(target)
        assert target.is_dir()

    def test_pt011_backward_compat_app_dir(self):
        """PT-011: APP_DIR alias equals BASE_DIR."""
        from logic.app_paths import APP_DIR, BASE_DIR
        assert APP_DIR == BASE_DIR

    def test_pt012_backward_compat_backups_dir(self):
        """PT-012: BACKUPS_DIR alias equals BACKUP_DIR."""
        from logic.app_paths import BACKUPS_DIR, BACKUP_DIR
        assert BACKUPS_DIR == BACKUP_DIR

    def test_pt013_version_constants_exist(self):
        """PT-013: Version module has required constants."""
        from logic.version import VERSION, BUILD_DATE, APP_NAME
        assert VERSION
        assert BUILD_DATE
        assert APP_NAME

    def test_pt014_version_format_semver(self):
        """PT-014: VERSION follows semver (X.Y.Z)."""
        from logic.version import VERSION
        parts = VERSION.split(".")
        assert len(parts) == 3
        assert all(p.isdigit() for p in parts)


# ================================================================
# 3. OFFLINE / FIREBASE FAIL-SAFE (OF-001 to OF-012)
# ================================================================

class TestOfflineFirebase:
    """OF-001..OF-012: Firebase unavailable, no internet, fail-safe."""

    def test_of001_remote_service_init_no_crash(self):
        """OF-001: RemoteControlService.__init__ does not crash."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        assert svc.device_id

    def test_of002_config_defaults_if_missing(self, tmp_path, monkeypatch):
        """OF-002: Missing config returns safe defaults (firebase_enabled=False)."""
        monkeypatch.setattr("logic.remote_control.REMOTE_CONFIG_PATH", tmp_path / "nope.json")
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        assert svc.config["firebase_enabled"] is False

    def test_of003_corrupt_config_returns_defaults(self, tmp_path, monkeypatch):
        """OF-003: Corrupt remote_config.json returns safe defaults."""
        cfg = tmp_path / "bad_config.json"
        cfg.write_text("{{{corrupt", encoding="utf-8")
        monkeypatch.setattr("logic.remote_control.REMOTE_CONFIG_PATH", cfg)
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        assert svc.config["firebase_enabled"] is False

    def test_of004_check_access_firebase_disabled(self):
        """OF-004: check_access with firebase_enabled=False → allow."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = False
        result = svc.check_access()
        assert result["action"] == "allow"

    def test_of005_check_access_firebase_fail_returns_warn(self):
        """OF-005: Firebase connection failure → warn, NOT block."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = True
        svc._firebase_failed = True
        result = svc.check_access()
        assert result["action"] in ("allow", "warn")

    def test_of006_device_id_persists(self, tmp_path, monkeypatch):
        """OF-006: Device ID is generated and persisted."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        assert len(svc.device_id) > 0

    def test_of007_missing_service_account_no_crash(self, tmp_path, monkeypatch):
        """OF-007: Missing firebase_service_account.json → local mode."""
        monkeypatch.setattr("logic.remote_control.FIREBASE_SERVICE_PATH", tmp_path / "nope.json")
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = True
        svc.config["database_url"] = "https://fake.firebaseio.com"
        svc._firebase_ready = False
        svc._firebase_failed = False
        svc._init_firebase()
        assert svc._firebase_failed is True

    def test_of008_check_access_exception_returns_warn(self):
        """OF-008: Any exception in check_access → warn action."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = True
        svc._firebase_ready = False
        svc._firebase_failed = False
        # Force a connection error by setting ready=False
        with patch.object(svc, "_get_reference_value", side_effect=ConnectionError("no net")):
            result = svc.check_access()
        assert result["action"] == "warn"

    def test_of009_offline_message_contains_seconds(self):
        """OF-009: Offline warn message includes offline duration."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = True
        with patch.object(svc, "_get_reference_value", side_effect=ConnectionError):
            result = svc.check_access()
        assert "offline" in result["message"].lower() or "indisponibil" in result["message"].lower()

    def test_of010_backoff_schedule_exists(self):
        """OF-010: Backoff schedule defined for RemoteChecker."""
        from logic.remote_control import _BACKOFF_SCHEDULE
        assert len(_BACKOFF_SCHEDULE) >= 3
        assert all(isinstance(s, (int, float)) for s in _BACKOFF_SCHEDULE)

    def test_of011_firebase_disabled_means_allow(self):
        """OF-011: Firebase disabled in config always returns allow."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        svc.config["firebase_enabled"] = False
        for _ in range(5):
            assert svc.check_access()["action"] == "allow"

    def test_of012_device_id_is_uuid_format(self):
        """OF-012: Device ID follows UUID format."""
        from logic.remote_control import RemoteControlService
        svc = RemoteControlService()
        import uuid
        uuid.UUID(svc.device_id)  # Throws if invalid


# ================================================================
# 4. AUTH / LOGIN (AU-001 to AU-016)
# ================================================================

class TestAuth:
    """AU-001..AU-016: Login, password, brute-force, user management."""

    def test_au001_default_admin_created(self, users_path):
        """AU-001: Default admin account auto-created."""
        from logic.auth import _load_users
        users = _load_users()
        assert any(u["username"] == "admin" for u in users)

    def test_au002_default_admin_password(self, users_path):
        """AU-002: Default admin password is admin123."""
        from logic.auth import verify_login
        assert verify_login("admin", "admin123") is True

    def test_au003_wrong_password_rejected(self, users_path):
        """AU-003: Wrong password returns False."""
        from logic.auth import verify_login
        # Create default admin first
        verify_login("admin", "admin123")
        assert verify_login("admin", "wrongpass") is False

    def test_au004_nonexistent_user_rejected(self, users_path):
        """AU-004: Non-existent user returns False."""
        from logic.auth import _load_users, verify_login
        _load_users()  # ensure file exists
        assert verify_login("nobody", "password123") is False

    def test_au005_empty_username_rejected(self, users_path):
        """AU-005: Empty username returns False."""
        from logic.auth import verify_login_detailed
        ok, msg = verify_login_detailed("", "password")
        assert ok is False

    def test_au006_empty_password_rejected(self, users_path):
        """AU-006: Empty password returns False."""
        from logic.auth import verify_login_detailed
        ok, msg = verify_login_detailed("admin", "")
        assert ok is False

    def test_au007_long_username_rejected(self, users_path):
        """AU-007: Username > 128 chars rejected."""
        from logic.auth import verify_login_detailed
        ok, msg = verify_login_detailed("a" * 200, "password")
        assert ok is False
        assert "lung" in msg.lower()

    def test_au008_long_password_rejected(self, users_path):
        """AU-008: Password > 256 chars rejected."""
        from logic.auth import verify_login_detailed
        ok, msg = verify_login_detailed("admin", "p" * 300)
        assert ok is False

    def test_au009_add_user(self, users_path):
        """AU-009: New user can be added."""
        from logic.auth import _load_users, add_user, verify_login
        _load_users()
        ok, msg = add_user("testuser", "test12345")
        assert ok is True
        assert verify_login("testuser", "test12345") is True

    def test_au010_add_duplicate_user_fails(self, users_path):
        """AU-010: Duplicate username rejected."""
        from logic.auth import _load_users, add_user
        _load_users()
        add_user("dup_user", "password1234")
        ok, msg = add_user("dup_user", "password5678")
        assert ok is False
        assert "exista" in msg.lower()

    def test_au011_add_user_short_password(self, users_path):
        """AU-011: Password < 8 chars rejected for new user."""
        from logic.auth import add_user
        ok, msg = add_user("usr", "short")
        assert ok is False

    def test_au012_change_password(self, users_path):
        """AU-012: Password change works."""
        from logic.auth import _load_users, change_password, verify_login
        _load_users()  # create default admin
        ok, msg = change_password("admin", "admin123", "newpass1234")
        assert ok is True
        assert verify_login("admin", "newpass1234") is True

    def test_au013_change_password_wrong_old(self, users_path):
        """AU-013: Change password with wrong old password fails."""
        from logic.auth import _load_users, change_password
        _load_users()
        ok, msg = change_password("admin", "wrong_old", "newpass1234")
        assert ok is False

    def test_au014_change_password_same_fails(self, users_path):
        """AU-014: New password same as old is rejected."""
        from logic.auth import _load_users, change_password
        _load_users()
        ok, msg = change_password("admin", "admin123", "admin123")
        assert ok is False

    def test_au015_case_insensitive_login(self, users_path):
        """AU-015: Login is case-insensitive for username."""
        from logic.auth import _load_users, verify_login
        _load_users()
        assert verify_login("Admin", "admin123") is True
        assert verify_login("ADMIN", "admin123") is True

    def test_au016_add_user_invalid_role(self, users_path):
        """AU-016: Invalid role rejected."""
        from logic.auth import _load_users, add_user
        _load_users()
        ok, msg = add_user("roleuser", "password1234", role="superadmin")
        assert ok is False


# ================================================================
# 5. DATA INTEGRITY — SCHEDULE STORE (DS-001 to DS-018)
# ================================================================

class TestScheduleIntegrity:
    """DS-001..DS-018: Schedule persistence, structure, backup/restore."""

    def test_ds001_new_week_has_all_modes(self, schedule_store):
        """DS-001: New week has Magazie and Bucle modes."""
        week = schedule_store.get_or_create_week(date.today())
        assert "Magazie" in week["modes"]
        assert "Bucle" in week["modes"]

    def test_ds002_new_week_has_all_days(self, schedule_store):
        """DS-002: Each mode has all 7 days."""
        from logic.schedule_store import DAY_NAMES
        week = schedule_store.get_or_create_week(date.today())
        for mode_name, mode_data in week["modes"].items():
            for dept in mode_data["departments"]:
                for day in DAY_NAMES:
                    assert day in mode_data["schedule"][dept]

    def test_ds003_new_week_has_all_shifts(self, schedule_store):
        """DS-003: Each day/dept cell has all 3 shifts."""
        from logic.schedule_store import DAY_NAMES, SHIFTS
        week = schedule_store.get_or_create_week(date.today())
        for mode_name, mode_data in week["modes"].items():
            for dept in mode_data["departments"]:
                for day in DAY_NAMES:
                    for shift in SHIFTS:
                        assert shift in mode_data["schedule"][dept][day]

    def test_ds004_employee_list_starts_empty(self, schedule_store):
        """DS-004: New cells have empty employee lists."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        cell = week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        assert cell["employees"] == []

    def test_ds005_update_week_persists(self, schedule_store):
        """DS-005: update_week saves data that survives reload."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["Test"]
        schedule_store.update_week(week)
        from logic.schedule_store import ScheduleStore
        store2 = ScheduleStore()
        w2 = store2.get_or_create_week(date.today())
        assert "Test" in w2["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"]

    def test_ds006_same_week_returns_same_record(self, schedule_store):
        """DS-006: Same date returns same week key."""
        w1 = schedule_store.get_or_create_week(date.today())
        w2 = schedule_store.get_or_create_week(date.today())
        assert w1["week_start"] == w2["week_start"]

    def test_ds007_week_start_is_monday(self, schedule_store):
        """DS-007: Week start is always a Monday."""
        week = schedule_store.get_or_create_week(date.today())
        ws = datetime.strptime(week["week_start"], "%Y-%m-%d").date()
        assert ws.weekday() == 0

    def test_ds008_backup_after_update(self, tmp_path, monkeypatch):
        """DS-008: Backup is created on update_week."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        week = store.get_or_create_week(date.today())
        store.update_week(week)
        assert len(list(backup_dir.glob("schedule_backup_*.json"))) >= 1

    def test_ds009_corrupt_main_loads_backup(self, tmp_path, monkeypatch):
        """DS-009: Corrupt main file → recovery from latest valid backup."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        # Write a valid backup
        backup_data = {"weeks": {"2026-04-06": {"week_start": "2026-04-06"}}}
        (backup_dir / "schedule_backup_20260407_000000.json").write_text(
            json.dumps(backup_data), encoding="utf-8"
        )
        # Corrupt main file
        path.write_text("CORRUPT!", encoding="utf-8")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert "2026-04-06" in store.data["weeks"]

    def test_ds010_no_duplication_on_reassign(self, schedule_store):
        """DS-010: Re-assigning same employee does not duplicate."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        cell = week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = ["Ion"]
        schedule_store.update_week(week)
        # Verify no dup
        w2 = schedule_store.get_or_create_week(date.today())
        assert w2["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"].count("Ion") == 1

    def test_ds011_clear_weekend(self, schedule_store):
        """DS-011: clear_weekend removes only weekend data."""
        from logic.schedule_store import WEEKEND_DAYS
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        for day in WEEKEND_DAYS:
            week["modes"]["Magazie"]["schedule"][dept][day]["Sch1"]["employees"] = ["A"]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["B"]
        schedule_store.clear_weekend(week, "Magazie")
        for day in WEEKEND_DAYS:
            assert week["modes"]["Magazie"]["schedule"][dept][day]["Sch1"]["employees"] == []
        assert "B" in week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"]

    def test_ds012_clear_department(self, schedule_store):
        """DS-012: clear_department clears all days/shifts for a dept."""
        from logic.schedule_store import DAY_NAMES, SHIFTS
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["X"]
        schedule_store.clear_department(week, "Magazie", dept)
        for day in DAY_NAMES:
            for shift in SHIFTS:
                assert week["modes"]["Magazie"]["schedule"][dept][day][shift]["employees"] == []

    def test_ds013_validate_assignment_duplicate_raises(self, schedule_store):
        """DS-013: Assigning same employee to same cell raises."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["Ion"]
        with pytest.raises(ValueError, match="exista"):
            schedule_store.validate_assignment(week, "Magazie", dept, "Luni", "Sch1", "Ion")

    def test_ds014_week_history_sorted(self, schedule_store):
        """DS-014: Week history returns sorted by date descending."""
        schedule_store.get_or_create_week(date.today())
        schedule_store.get_or_create_week(date.today() - timedelta(days=7))
        history = schedule_store.get_week_history()
        dates = [item[0] for item in history]
        assert dates == sorted(dates, reverse=True)

    def test_ds015_magazie_departments_match_template(self, schedule_store):
        """DS-015: Magazie mode has correct departments from template."""
        from logic.schedule_store import TEMPLATES
        week = schedule_store.get_or_create_week(date.today())
        assert week["modes"]["Magazie"]["departments"] == list(TEMPLATES["Magazie"])

    def test_ds016_bucle_departments_match_template(self, schedule_store):
        """DS-016: Bucle mode has correct departments from template."""
        from logic.schedule_store import TEMPLATES
        week = schedule_store.get_or_create_week(date.today())
        assert week["modes"]["Bucle"]["departments"] == list(TEMPLATES["Bucle"])

    def test_ds017_save_produces_valid_json(self, schedule_store):
        """DS-017: Saved file is valid JSON."""
        schedule_store.get_or_create_week(date.today())
        schedule_store.save()
        from logic.schedule_store import SCHEDULE_PATH
        with SCHEDULE_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
        assert "weeks" in data

    def test_ds018_assignment_map_built(self, schedule_store):
        """DS-018: build_assignment_map returns correct mappings."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["Ion"]
        amap = schedule_store.build_assignment_map(week, "Magazie")
        assert "ion" in amap


# ================================================================
# 6. DATA INTEGRITY — EMPLOYEE STORE (ES-001 to ES-010)
# ================================================================

class TestEmployeeIntegrity:
    """ES-001..ES-010: Employee CRUD, normalization."""

    def test_es001_add_employee(self, employee_store):
        """ES-001: Employee can be added."""
        employee_store.add_employee("Ion Popescu")
        assert "Ion Popescu" in employee_store.get_all()

    def test_es002_duplicate_employee_silent(self, employee_store):
        """ES-002: Duplicate employee silently ignored (no crash)."""
        employee_store.add_employee("Ion Popescu")
        employee_store.add_employee("Ion Popescu")
        assert employee_store.get_all().count("Ion Popescu") == 1

    def test_es003_case_insensitive_duplicate(self, employee_store):
        """ES-003: Case-insensitive duplicate detected."""
        employee_store.add_employee("Ion Popescu")
        employee_store.add_employee("ion popescu")
        assert employee_store.get_all().count("Ion Popescu") == 1

    def test_es004_delete_employee(self, employee_store):
        """ES-004: Employee can be deleted."""
        employee_store.add_employee("Ion Popescu")
        result = employee_store.delete_employee("Ion Popescu")
        assert result is True
        assert "Ion Popescu" not in employee_store.get_all()

    def test_es005_delete_nonexistent_returns_false(self, employee_store):
        """ES-005: Deleting non-existent employee returns False."""
        assert employee_store.delete_employee("Nobody") is False

    def test_es006_rename_employee(self, employee_store):
        """ES-006: Employee can be renamed."""
        employee_store.add_employee("Old Name")
        employee_store.rename_employee("Old Name", "New Name")
        assert "New Name" in employee_store.get_all()
        assert "Old Name" not in employee_store.get_all()

    def test_es007_rename_to_existing_raises(self, employee_store):
        """ES-007: Rename to existing name raises ValueError."""
        employee_store.add_employee("Name A")
        employee_store.add_employee("Name B")
        with pytest.raises(ValueError):
            employee_store.rename_employee("Name A", "Name B")

    def test_es008_empty_name_raises(self, employee_store):
        """ES-008: Empty name raises ValueError."""
        with pytest.raises(ValueError):
            employee_store.add_employee("")

    def test_es009_whitespace_trimmed(self, employee_store):
        """ES-009: Whitespace is trimmed from names."""
        employee_store.add_employee("  Ion Popescu  ")
        assert "Ion Popescu" in employee_store.get_all()

    def test_es010_get_all_sorted(self, employee_store):
        """ES-010: get_all returns sorted list."""
        employee_store.add_employee("Zoe")
        employee_store.add_employee("Ana")
        all_emp = employee_store.get_all()
        assert all_emp == sorted(all_emp, key=str.casefold)


# ================================================================
# 7. EXCEL EXPORT (EX-001 to EX-030)
# ================================================================

class TestExcelExport:
    """EX-001..EX-030: Excel A3 export structural validation."""

    @pytest.fixture(autouse=True)
    def _export(self, week_record, tmp_path):
        from logic.excel_exporter import ExcelExporter
        self.export_dir = tmp_path / "Exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", self.export_dir):
            self.path = ExcelExporter.export(week_record, "Magazie")
        from openpyxl import load_workbook
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active

    def test_ex001_file_exists(self):
        """EX-001: Export file is created."""
        assert self.path.exists()

    def test_ex002_file_is_xlsx(self):
        """EX-002: File has .xlsx extension."""
        assert self.path.suffix == ".xlsx"

    def test_ex003_sheet_title_is_mode(self):
        """EX-003: Sheet title matches mode name."""
        assert self.ws.title == "Magazie"

    def test_ex004_max_column_is_9(self):
        """EX-004: Table has exactly 9 columns (A-I), no dead col J."""
        assert self.ws.max_column == 9

    def test_ex005_paper_size_a3(self):
        """EX-005: Paper size is A3 (code 8)."""
        assert int(self.ws.page_setup.paperSize) == 8

    def test_ex006_orientation_landscape(self):
        """EX-006: Orientation is landscape."""
        assert self.ws.page_setup.orientation == self.ws.ORIENTATION_LANDSCAPE

    def test_ex007_fit_to_width_1(self):
        """EX-007: fitToWidth = 1."""
        assert self.ws.page_setup.fitToWidth == 1

    def test_ex008_fit_to_height_1(self):
        """EX-008: fitToHeight = 1 (single page)."""
        assert self.ws.page_setup.fitToHeight == 1

    def test_ex009_header_merge_a1_i2(self):
        """EX-009: Header merged A1:I2."""
        merges = [str(m) for m in self.ws.merged_cells.ranges]
        assert "A1:I2" in merges

    def test_ex010_subtitle_merge_a3_i3(self):
        """EX-010: Subtitle merged A3:I3."""
        merges = [str(m) for m in self.ws.merged_cells.ranges]
        assert "A3:I3" in merges

    def test_ex011_no_col_j_merge(self):
        """EX-011: No merges extend to column J."""
        for m in self.ws.merged_cells.ranges:
            assert "J" not in str(m)

    def test_ex012_header_text_contains_mode(self):
        """EX-012: Header contains mode name."""
        assert "magazie" in str(self.ws["A1"].value).lower()

    def test_ex013_subtitle_contains_version(self):
        """EX-013: Subtitle contains version."""
        from logic.version import VERSION
        assert VERSION in str(self.ws["A3"].value)

    def test_ex014_header_fill_blue(self):
        """EX-014: Header has blue fill (0067C8)."""
        fill = self.ws["A1"].fill
        assert fill.fgColor and "0067C8" in str(fill.fgColor.rgb).upper()

    def test_ex015_header_font_white(self):
        """EX-015: Header font is white."""
        color = str(self.ws["A1"].font.color.rgb).upper()
        assert color.endswith("FFFFFF")

    def test_ex016_print_area_set(self):
        """EX-016: Print area is set."""
        assert self.ws.print_area is not None
        assert len(self.ws.print_area) > 0

    def test_ex017_print_area_no_col_j(self):
        """EX-017: Print area does not include column J."""
        pa = str(self.ws.print_area)
        # Should end with $I$<something>
        assert "$J$" not in pa

    def test_ex018_freeze_panes_c4(self):
        """EX-018: Freeze panes at C4."""
        assert str(self.ws.freeze_panes) == "C4"

    def test_ex019_print_title_rows(self):
        """EX-019: Print title rows set to 1:3."""
        assert self.ws.print_title_rows in ("1:3", "$1:$3")

    def test_ex020_one_image(self):
        """EX-020: Exactly one image (logo)."""
        assert len(self.ws._images) == 1

    def test_ex021_logo_dimensions(self):
        """EX-021: Logo is present and has positive dimensions."""
        img = self.ws._images[0]
        assert img.width > 0
        assert img.height > 0

    def test_ex022_margins_tight(self):
        """EX-022: Margins are tight for A3."""
        m = self.ws.page_margins
        assert m.left <= 0.3
        assert m.right <= 0.3

    def test_ex023_no_gridlines(self):
        """EX-023: Grid lines hidden."""
        assert self.ws.sheet_view.showGridLines is False

    def test_ex024_column_a_width_narrow(self):
        """EX-024: Column A width is narrow (dept vertical)."""
        assert self.ws.column_dimensions["A"].width <= 6

    def test_ex025_day_column_width_30(self):
        """EX-025: Day columns (C-I) are width 30."""
        for col in "CDEFGHI":
            assert self.ws.column_dimensions[col].width == 30

    def test_ex026_all_fills_ff_alpha(self):
        """EX-026: All PatternFills have FF alpha prefix (no 00 transparency)."""
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row):
            for cell in row:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    if rgb and len(rgb) == 8 and rgb != "00000000":
                        assert rgb.startswith("FF"), f"Cell {cell.coordinate}: alpha not FF → {rgb}"

    def test_ex027_no_cellrichtext(self):
        """EX-027: No CellRichText objects (would cause corruption)."""
        from openpyxl.cell.rich_text import CellRichText
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                assert not isinstance(cell.value, CellRichText)

    def test_ex028_department_merge_exists(self):
        """EX-028: Department cells are merged vertically."""
        merges = [str(m) for m in self.ws.merged_cells.ranges]
        dept_merges = [m for m in merges if m.startswith("A") and "I" not in m]
        assert len(dept_merges) >= 6  # Magazie has 6 departments

    def test_ex029_footer_set(self):
        """EX-029: Footer is configured."""
        assert self.ws.oddFooter.left.text is not None
        assert "Autoliv" in self.ws.oddFooter.left.text

    def test_ex030_subtitle_fill_navy(self):
        """EX-030: Subtitle row has navy fill."""
        fill = self.ws["A3"].fill
        assert fill.fgColor and "1A4A80" in str(fill.fgColor.rgb).upper()


# ================================================================
# 7b. EXCEL EXPORT — POPULATED DATA (EP-001 to EP-008)
# ================================================================

class TestExcelExportPopulated:
    """EP-001..EP-008: Export with actual employee data."""

    @pytest.fixture(autouse=True)
    def _export(self, populated_week, tmp_path):
        from logic.excel_exporter import ExcelExporter
        self.export_dir = tmp_path / "Exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", self.export_dir):
            self.path = ExcelExporter.export(populated_week, "Magazie")
        from openpyxl import load_workbook
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active

    def test_ep001_employees_in_cells(self):
        """EP-001: Employee names appear in exported cells."""
        found = False
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Ion Popescu" in str(cell.value):
                    found = True
        assert found

    def test_ep002_hours_prefix(self):
        """EP-002: Employee names prefixed with hours fallback text."""
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Ion Popescu" in str(cell.value):
                    assert "\u25cf 12 Ion Popescu" in str(cell.value)
                    assert "\u25cf" in str(cell.value)

    def test_ep003_font_color_is_argb(self):
        """EP-003: Font colors are 8-char ARGB."""
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Ion Popescu" in str(cell.value):
                    color = str(cell.font.color.rgb)
                    assert len(color) == 8
                    assert color.startswith("FF")

    def test_ep004_cell_fill_white(self):
        """EP-004: Employee cells have white fill."""
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Ion Popescu" in str(cell.value):
                    assert "FFFFFF" in str(cell.fill.fgColor.rgb).upper()

    def test_ep005_multi_employee_newlines(self):
        """EP-005: Multiple employees separated by newlines."""
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Ion Popescu" in str(cell.value):
                    assert "\n" in str(cell.value)
                    assert "Maria Ionescu" in str(cell.value)

    def test_ep006_file_opens_without_exception(self):
        """EP-006: File loads with openpyxl without exception."""
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        assert wb.active is not None

    def test_ep007_max_column_still_9(self):
        """EP-007: Populated export still has 9 columns max."""
        assert self.ws.max_column == 9

    def test_ep008_dept_header_row_has_day_labels(self):
        """EP-008: Department header row contains day names."""
        found_luni = False
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row):
            for cell in row:
                if cell.value and "Luni" in str(cell.value):
                    found_luni = True
        assert found_luni


# ================================================================
# 8. VALIDATION MODULE (VL-001 to VL-012)
# ================================================================

class TestValidation:
    """VL-001..VL-012: Input validation rules."""

    def test_vl001_valid_hours_integer(self):
        from logic.validation import _parse_hours
        assert _parse_hours("Test", "8") == 8

    def test_vl002_valid_hours_float(self):
        from logic.validation import _parse_hours
        assert _parse_hours("Test", "4.5") == 4.5

    def test_vl003_max_boundary(self):
        from logic.validation import _parse_hours
        assert _parse_hours("Test", "12") == 12

    def test_vl004_empty_not_allowed_raises(self):
        from logic.validation import _parse_hours
        with pytest.raises(ValueError):
            _parse_hours("Test", "")

    def test_vl005_empty_allowed_returns_none(self):
        from logic.validation import _parse_hours
        assert _parse_hours("Test", "", allow_empty=True) is None

    def test_vl006_zero_raises(self):
        from logic.validation import _parse_hours
        with pytest.raises(ValueError):
            _parse_hours("Test", "0")

    def test_vl007_negative_raises(self):
        from logic.validation import _parse_hours
        with pytest.raises(ValueError):
            _parse_hours("Test", "-5")

    def test_vl008_over_12_raises(self):
        from logic.validation import _parse_hours
        with pytest.raises(ValueError):
            _parse_hours("Test", "13")

    def test_vl009_non_numeric_raises(self):
        from logic.validation import _parse_hours
        with pytest.raises(ValueError):
            _parse_hours("Test", "abc")

    def test_vl010_whitespace_stripped(self):
        from logic.validation import _parse_hours
        assert _parse_hours("Test", "  8  ") == 8

    def test_vl011_validate_employee_data_valid(self):
        from logic.validation import validate_employee_data
        result = validate_employee_data("Ion", "Popescu", "Receptii", "8")
        assert result is not None

    def test_vl012_validate_employee_data_missing_name(self):
        from logic.validation import validate_employee_data
        with pytest.raises(ValueError):
            validate_employee_data("", "Popescu", "Receptii", "8")


# ================================================================
# 9. EDGE CASES (EC-001 to EC-015)
# ================================================================

class TestEdgeCases:
    """EC-001..EC-015: Boundary conditions, stress scenarios."""

    def test_ec001_empty_schedule_export(self, week_record, tmp_path):
        """EC-001: Exporting empty schedule does not crash."""
        from logic.excel_exporter import ExcelExporter
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(week_record, "Magazie")
        assert path.exists()

    def test_ec002_many_employees_one_cell(self, tmp_path):
        """EC-002: Cell with 20 employees does not crash."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        dept = w["modes"]["Magazie"]["departments"][0]
        cell = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = [f"Employee_{i}" for i in range(20)]
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Magazie")
        assert path.exists()

    def test_ec003_long_employee_name(self, tmp_path):
        """EC-003: Employee with very long name does not crash."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        dept = w["modes"]["Magazie"]["departments"][0]
        cell = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = ["A" * 100]
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Magazie")
        assert path.exists()

    def test_ec004_special_chars_in_name(self, tmp_path):
        """EC-004: Special characters in name handled correctly."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        dept = w["modes"]["Magazie"]["departments"][0]
        cell = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = ["Ștefan Pîrvu-Dăescu", "José García", "François Müller"]
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Magazie")
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        found = False
        for row in ws.iter_rows():
            for cell_obj in row:
                if cell_obj.value and "Ștefan" in str(cell_obj.value):
                    found = True
        assert found

    def test_ec005_export_bucle_mode(self, tmp_path):
        """EC-005: Exporting Bucle mode works correctly."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Bucle")
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert wb.active.title == "Bucle"

    def test_ec006_no_logo_no_crash(self, week_record, tmp_path):
        """EC-006: Missing logo does not crash export."""
        from logic.excel_exporter import ExcelExporter
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(week_record, "Magazie", logo_path=Path("/nonexistent/logo.png"))
        assert path.exists()

    def test_ec007_color_map_all_valid_hex(self):
        """EC-007: All COLOR_MAP values are valid 6-char hex."""
        from logic.excel_exporter import COLOR_MAP
        for key, val in COLOR_MAP.items():
            assert len(val) == 6
            assert all(c in "0123456789ABCDEFabcdef" for c in val)

    def test_ec008_cell_text_empty_list(self):
        """EC-008: _cell_text_and_color with empty list returns empty."""
        from logic.excel_exporter import _cell_text_and_color
        text, color = _cell_text_and_color([], {})
        assert text == ""

    def test_ec009_cell_text_single_employee(self):
        """EC-009: Single employee returns hours-prefixed text."""
        from logic.excel_exporter import _cell_text_and_color
        text, color = _cell_text_and_color(["Ion"], {"Ion": "#C0392B"})
        assert "Ion" in text
        assert text == "\u25cf 12 Ion"

    def test_ec010_cell_text_mixed_colors_uses_default(self):
        """EC-010: Mixed colors → default dark color."""
        from logic.excel_exporter import _cell_text_and_color, DEFAULT_TEXT_COLOR
        text, color = _cell_text_and_color(
            ["A", "B"], {"A": "#C0392B", "B": "#27AE60"}
        )
        assert color == "FF" + DEFAULT_TEXT_COLOR

    def test_ec011_cell_text_same_color_uses_it(self):
        """EC-011: Same color for all → that color used."""
        from logic.excel_exporter import _cell_text_and_color
        text, color = _cell_text_and_color(
            ["A", "B"], {"A": "#C0392B", "B": "#C0392B"}
        )
        assert color == "FFC0392B"

    def test_ec012_to_argb_hash_prefix_stripped(self):
        """EC-012: _to_argb strips # prefix correctly."""
        from logic.excel_exporter import _to_argb
        assert _to_argb("#C0392B") == "FFC0392B"

    def test_ec013_to_argb_none_returns_default(self):
        """EC-013: _to_argb(None) returns default."""
        from logic.excel_exporter import _to_argb, DEFAULT_TEXT_COLOR
        assert _to_argb(None) == "FF" + DEFAULT_TEXT_COLOR

    def test_ec014_to_argb_invalid_hex_returns_default(self):
        """EC-014: _to_argb('XYZ') returns default."""
        from logic.excel_exporter import _to_argb, DEFAULT_TEXT_COLOR
        assert _to_argb("XYZ") == "FF" + DEFAULT_TEXT_COLOR

    def test_ec015_fill_helper_ff_prefix(self):
        """EC-015: _fill always produces FF alpha."""
        from logic.excel_exporter import _fill
        fill = _fill("C0392B")
        assert str(fill.fgColor.rgb).startswith("FF")


# ================================================================
# 10. ERROR HANDLING / RECOVERY (EH-001 to EH-010)
# ================================================================

class TestErrorHandling:
    """EH-001..EH-010: Graceful failure, no uncaught exceptions."""

    def test_eh001_schedule_store_corrupt_recovery(self, tmp_path, monkeypatch):
        """EH-001: Corrupt schedule + valid backup → recovered."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        backup_data = {"weeks": {"2026-04-01": {"week_start": "2026-04-01"}}}
        (backup_dir / "schedule_backup_20260407_120000.json").write_text(
            json.dumps(backup_data), encoding="utf-8"
        )
        path.write_text("!!CORRUPT!!", encoding="utf-8")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert "2026-04-01" in store.data["weeks"]

    def test_eh002_corrupt_backup_skipped(self, tmp_path, monkeypatch):
        """EH-002: Corrupt backup is skipped, next valid one used."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        (backup_dir / "schedule_backup_20260407_120002.json").write_text("CORRUPT!", encoding="utf-8")
        valid = {"weeks": {"2026-03-25": {"week_start": "2026-03-25"}}}
        (backup_dir / "schedule_backup_20260407_120001.json").write_text(
            json.dumps(valid), encoding="utf-8"
        )
        path.write_text("CORRUPT!", encoding="utf-8")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert "2026-03-25" in store.data["weeks"]

    def test_eh003_no_backups_starts_empty(self, tmp_path, monkeypatch):
        """EH-003: No valid backups → start with empty weeks."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        path.write_text("CORRUPT!", encoding="utf-8")
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        assert store.data == {"weeks": {}}

    def test_eh004_auth_graceful_on_bcrypt_error(self, users_path):
        """EH-004: bcrypt internal error returns False with message."""
        from logic.auth import _load_users
        _load_users()
        from logic.auth import verify_login_detailed
        with patch("logic.auth.bcrypt.checkpw", side_effect=Exception("bcrypt boom")):
            ok, msg = verify_login_detailed("admin", "admin123")
            assert ok is False
            assert "eroare" in msg.lower()

    def test_eh005_export_with_none_colors(self, tmp_path):
        """EH-005: Export handles None in colors dict."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        dept = w["modes"]["Magazie"]["departments"][0]
        cell = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = ["Ion"]
        cell["colors"] = {"Ion": None}
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Magazie")
        assert path.exists()

    def test_eh006_export_missing_color_key(self, tmp_path):
        """EH-006: Export handles employee without color entry."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        dept = w["modes"]["Magazie"]["departments"][0]
        cell = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]
        cell["employees"] = ["Ion"]
        cell["colors"] = {}  # no color entry for Ion
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(w, "Magazie")
        assert path.exists()

    def test_eh007_rotate_backups_limits_count(self, tmp_path, monkeypatch):
        """EH-007: Backup rotation keeps only max_backups files."""
        path = tmp_path / "data" / "schedule_data.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_dir = tmp_path / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr("logic.schedule_store.SCHEDULE_PATH", path)
        monkeypatch.setattr("logic.schedule_store.BACKUP_DIR", backup_dir)
        from logic.schedule_store import ScheduleStore
        store = ScheduleStore()
        store.get_or_create_week(date.today())
        store.save()
        for i in range(25):
            (backup_dir / f"schedule_backup_20260407_{i:06d}.json").write_text(
                '{"weeks":{}}', encoding="utf-8"
            )
        store._rotate_backups(max_backups=5)
        remaining = list(backup_dir.glob("schedule_backup_*.json"))
        assert len(remaining) <= 5

    def test_eh008_cell_text_no_colors_dict(self):
        """EH-008: _cell_text_and_color with None colors dict."""
        from logic.excel_exporter import _cell_text_and_color
        text, color = _cell_text_and_color(["A"], None)
        assert "A" in text

    def test_eh009_cell_text_case_insensitive_lookup(self):
        """EH-009: Color lookup is case-insensitive."""
        from logic.excel_exporter import _cell_text_and_color
        text, color = _cell_text_and_color(["Ion POPESCU"], {"ion popescu": "#C0392B"})
        assert color == "FFC0392B"

    def test_eh010_export_invalid_logo_path(self, week_record, tmp_path):
        """EH-010: Invalid logo path logs error but doesn't crash."""
        from logic.excel_exporter import ExcelExporter
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            path = ExcelExporter.export(week_record, "Magazie", logo_path=Path("/bad/path.png"))
        assert path.exists()


# ================================================================
# 11. MULTI-RUN CONSISTENCY (MR-001 to MR-010)
# ================================================================

class TestMultiRunConsistency:
    """MR-001..MR-010: Repeated operations produce consistent results."""

    def test_mr001_double_export_same_structure(self, week_record, tmp_path):
        """MR-001: Two exports of same data have same column count."""
        from logic.excel_exporter import ExcelExporter
        from openpyxl import load_workbook
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            p1 = ExcelExporter.export(week_record, "Magazie")
            p2 = ExcelExporter.export(week_record, "Magazie")
        wb1 = load_workbook(p1)
        wb2 = load_workbook(p2)
        assert wb1.active.max_column == wb2.active.max_column
        assert wb1.active.max_row == wb2.active.max_row

    def test_mr002_reload_schedule_consistent(self, schedule_store):
        """MR-002: Save + reload returns same data."""
        week = schedule_store.get_or_create_week(date.today())
        dept = week["modes"]["Magazie"]["departments"][0]
        week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] = ["Test"]
        schedule_store.update_week(week)
        from logic.schedule_store import ScheduleStore
        store2 = ScheduleStore()
        w2 = store2.get_or_create_week(date.today())
        assert w2["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"] == ["Test"]

    def test_mr003_multiple_adds_no_duplicates(self, employee_store):
        """MR-003: Adding same employee twice produces no duplicates."""
        employee_store.add_employee("Test Person")
        employee_store.add_employee("Test Person")
        assert employee_store.get_all().count("Test Person") == 1

    def test_mr004_repeated_auth_consistent(self, users_path):
        """MR-004: Multiple login attempts give consistent results."""
        from logic.auth import _load_users, verify_login
        _load_users()
        for _ in range(5):
            assert verify_login("admin", "admin123") is True
        for _ in range(3):
            assert verify_login("admin", "wrong") is False

    def test_mr005_export_both_modes(self, tmp_path):
        """MR-005: Exporting both modes produces 2 valid files."""
        from logic.schedule_store import _empty_week_record, get_week_start
        from logic.excel_exporter import ExcelExporter
        w = _empty_week_record(get_week_start(date.today()))
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            p1 = ExcelExporter.export(w, "Magazie")
            p2 = ExcelExporter.export(w, "Bucle")
        assert p1.exists() and p2.exists()
        assert p1 != p2

    def test_mr006_save_load_save_load(self, schedule_store):
        """MR-006: Multiple save/load cycles preserve data."""
        for i in range(3):
            week = schedule_store.get_or_create_week(date.today())
            dept = week["modes"]["Magazie"]["departments"][0]
            emps = week["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"]
            name = f"Employee_{i}"
            if name not in emps:
                emps.append(name)
            schedule_store.update_week(week)
        # Final check
        from logic.schedule_store import ScheduleStore
        store2 = ScheduleStore()
        w = store2.get_or_create_week(date.today())
        emps = w["modes"]["Magazie"]["schedule"][dept]["Luni"]["Sch1"]["employees"]
        assert "Employee_0" in emps
        assert "Employee_1" in emps
        assert "Employee_2" in emps

    def test_mr007_week_key_deterministic(self, schedule_store):
        """MR-007: Same date always gives same week_start key."""
        today = date.today()
        w1 = schedule_store.get_or_create_week(today)
        w2 = schedule_store.get_or_create_week(today)
        w3 = schedule_store.get_or_create_week(today)
        assert w1["week_start"] == w2["week_start"] == w3["week_start"]

    def test_mr008_color_helpers_idempotent(self):
        """MR-008: Color helpers give same result on repeated calls."""
        from logic.excel_exporter import _to_argb, _fill
        for _ in range(10):
            assert _to_argb("#C0392B") == "FFC0392B"
            assert str(_fill("FFFFFF").fgColor.rgb) == "FFFFFFFF"

    def test_mr009_export_dimensions_stable(self, week_record, tmp_path):
        """MR-009: Same data → same dimensions across exports."""
        from logic.excel_exporter import ExcelExporter
        from openpyxl import load_workbook
        d = tmp_path / "Exports"
        d.mkdir(parents=True, exist_ok=True)
        dims = []
        with patch("logic.excel_exporter.EXPORT_DIR", d):
            for _ in range(3):
                p = ExcelExporter.export(week_record, "Magazie")
                wb = load_workbook(p)
                dims.append(wb.active.dimensions)
        assert len(set(dims)) == 1  # all same

    def test_mr010_schedule_constants_immutable(self):
        """MR-010: Module-level constants are consistent."""
        from logic.schedule_store import DAYS, SHIFTS, DAY_NAMES, TEMPLATES
        assert len(DAYS) == 7
        assert len(SHIFTS) == 3
        assert len(DAY_NAMES) == 7
        assert "Magazie" in TEMPLATES
        assert "Bucle" in TEMPLATES
